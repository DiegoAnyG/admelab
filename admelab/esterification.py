"""
admelab.esterification
======================
Targeted esterification of a carboxylic acid with a library of alcohols.

Designed for "one acid core x N alcohols" series (e.g. a carboxylic scaffold x
inventory alcohols), with **absolute priority on chemical correctness**:

  - Real SMARTS reaction (not text manipulation), so the alcohol
    **stereochemistry is preserved** in the ester.
  - Classification of each OH (primary / secondary / tertiary / phenolic) and a
    **Fischer viability flag**.
  - Explicit **regioselectivity** when the alcohol has several OH groups: all
    regioisomers or only the preferred one, depending on policy.
  - **Validation of each product**: exact expected atomic formula
    (alcohol + acid - H2O), presence of the ester group, preservation of chiral
    centers and RDKit sanitization.
  - **Audit** against a reference SMILES column (if you already have a previous
    list), comparing by InChIKey.

This module is the specialization of `admelab.reactions` for the acid + alcohol
case; for other reactions (e.g. amidation) use `admelab.reactions.react_library`.

Main function: `esterify_library(...)` -> DataFrame.
"""
from __future__ import annotations

from collections import Counter
from typing import Optional, Sequence

import pandas as pd
from rdkit import Chem
from rdkit import RDLogger
from rdkit.Chem import AllChem, Descriptors, rdMolDescriptors as rdmd

from .reactions import ESTERIFICATION_SMARTS  # single source of the SMARTS

RDLogger.DisableLog("rdApp.*")


# --- hydroxyl classification patterns --------------------------------------
_PATTERNS = {
    "phenolic":  Chem.MolFromSmarts("[OX2H][c]"),
    "tertiary":  Chem.MolFromSmarts("[OX2H][CX4]([#6])([#6])[#6]"),
    "secondary": Chem.MolFromSmarts("[OX2H][CX4H1]([#6])[#6]"),
    "primary":   Chem.MolFromSmarts("[OX2H][CX4H2][#6]"),
    "methanol":  Chem.MolFromSmarts("[OX2H][CX4H3]"),
}
_ESTER_PATTERN = Chem.MolFromSmarts("[CX3](=[OX1])[OX2][#6]")
_COOH_PATTERN = Chem.MolFromSmarts("[CX3](=[OX1])[OX2H1]")

# Rough viability of Fischer esterification by OH type.
FISCHER_VIABILITY = {
    "primary":   "good",
    "methanol":  "good",
    "secondary": "moderate",
    "tertiary":  "unfavorable (E1/SN1; use acyl chloride or Steglich)",
    "phenolic":  "difficult (weakly nucleophilic phenol; use acyl chloride/DMAP)",
    "unknown":   "unknown",
}

# Regioselectivity preference when several OH are present (higher = reacts first).
_PRIORITY = {"primary": 4, "methanol": 4, "secondary": 3, "phenolic": 2,
             "tertiary": 1, "unknown": 0}


# ---------------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------------
def _atom_counts(mol: Chem.Mol) -> Counter:
    """Exact atomic count including hydrogens."""
    return Counter(a.GetSymbol() for a in Chem.AddHs(mol).GetAtoms())


def classify_hydroxyls(mol: Chem.Mol) -> list[dict]:
    """Classify each OH of the alcohol. Returns [{'o_idx', 'type', 'viability'}]."""
    if mol is None:
        return []
    found: dict[int, str] = {}
    # Order matters: most specific first (tertiary before secondary).
    for kind in ("phenolic", "tertiary", "secondary", "primary", "methanol"):
        patt = _PATTERNS[kind]
        if patt is None:
            continue
        for match in mol.GetSubstructMatches(patt):
            o_idx = match[0]
            found.setdefault(o_idx, kind)
    # OH not classified (e.g. attached to a non-aromatic sp2 C)
    for atom in mol.GetAtoms():
        if atom.GetSymbol() == "O" and atom.GetTotalNumHs() >= 1 and atom.GetDegree() == 1:
            found.setdefault(atom.GetIdx(), "unknown")
    return [{"o_idx": i, "type": t, "viability": FISCHER_VIABILITY.get(t, "unknown")}
            for i, t in sorted(found.items())]


def _oh_type_for_product(alcohol: Chem.Mol, product: Chem.Mol) -> str:
    """Infer which OH type reacted by comparing the remaining OH groups."""
    before = classify_hydroxyls(alcohol)
    after_types = Counter(h["type"] for h in classify_hydroxyls(product))
    before_types = Counter(h["type"] for h in before)
    diff = before_types - after_types
    return next(iter(diff), "unknown") if diff else "unknown"


# ---------------------------------------------------------------------------
# Reaction
# ---------------------------------------------------------------------------
def esterify(
    acid_smiles: str,
    alcohol_smiles: str,
    policy: str = "all",
    smarts: str = ESTERIFICATION_SMARTS,
) -> list[dict]:
    """Esterify an alcohol with the acid. Returns a list of products.

    policy:
      - "all"       : all regioisomers (one product per OH).
      - "preferred" : only the most Fischer-favorable OH
                      (primary > secondary > phenolic > tertiary).

    Each product: {'smiles', 'inchikey', 'oh_type', 'fischer_viability', 'n_regioisomers'}
    """
    acid = Chem.MolFromSmiles(acid_smiles)
    alcohol = Chem.MolFromSmiles(alcohol_smiles)
    if acid is None or alcohol is None:
        return []
    if not acid.HasSubstructMatch(_COOH_PATTERN):
        raise ValueError("The acid SMILES does not contain a -COOH group.")

    rxn = AllChem.ReactionFromSmarts(smarts)
    outcomes = rxn.RunReactants((acid, alcohol))

    uniq: dict[str, Chem.Mol] = {}
    for tup in outcomes:
        prod = tup[0]
        try:
            Chem.SanitizeMol(prod)
            # Reassign stereochemistry from the resulting structure.
            Chem.AssignStereochemistry(prod, cleanIt=True, force=True)
            uniq[Chem.MolToSmiles(prod)] = prod
        except Exception:
            continue

    products = []
    for smi, mol in uniq.items():
        oh_type = _oh_type_for_product(alcohol, mol)
        products.append({
            "smiles": smi,
            "inchikey": Chem.MolToInchiKey(mol),
            "oh_type": oh_type,
            "fischer_viability": FISCHER_VIABILITY.get(oh_type, "unknown"),
            "n_regioisomers": len(uniq),
        })

    if policy == "preferred" and len(products) > 1:
        products.sort(key=lambda p: _PRIORITY.get(p["oh_type"], 0), reverse=True)
        products = products[:1]
    return products


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------
def validate_ester(acid_smiles: str, alcohol_smiles: str, product_smiles: str) -> dict:
    """Check that the product is really the expected ester.

    Verifies: parsing, exact atomic formula (alcohol + acid - H2O), presence of
    the ester group, and that no chiral centers of the alcohol were lost.
    """
    acid = Chem.MolFromSmiles(acid_smiles)
    alc = Chem.MolFromSmiles(alcohol_smiles)
    prod = Chem.MolFromSmiles(product_smiles) if product_smiles else None
    res = {"valid": False, "formula_ok": False, "has_ester": False,
           "stereo_ok": False, "formula": None, "reason": ""}
    if prod is None:
        res["reason"] = "product did not parse"
        return res
    if acid is None or alc is None:
        res["reason"] = "reactant did not parse"
        return res

    expected = _atom_counts(alc) + _atom_counts(acid)
    expected["H"] -= 2
    expected["O"] -= 1
    expected = Counter({k: v for k, v in expected.items() if v > 0})
    obtained = _atom_counts(prod)
    res["formula"] = rdmd.CalcMolFormula(prod)
    res["formula_ok"] = (expected == obtained)

    res["has_ester"] = prod.HasSubstructMatch(_ESTER_PATTERN)

    n_alc = len(Chem.FindMolChiralCenters(alc, useLegacyImplementation=False,
                                          includeUnassigned=False))
    n_prod = len(Chem.FindMolChiralCenters(prod, useLegacyImplementation=False,
                                           includeUnassigned=False))
    res["stereo_ok"] = (n_prod >= n_alc)

    reasons = []
    if not res["formula_ok"]:
        reasons.append(f"formula {res['formula']} != expected")
    if not res["has_ester"]:
        reasons.append("no ester group")
    if not res["stereo_ok"]:
        reasons.append(f"chiral centers lost ({n_alc}->{n_prod})")
    res["reason"] = "; ".join(reasons)
    res["valid"] = res["formula_ok"] and res["has_ester"] and res["stereo_ok"]
    return res


# ---------------------------------------------------------------------------
# Facade: full library
# ---------------------------------------------------------------------------
def esterify_library(
    acid_smiles: str,
    alcohols: pd.DataFrame | Sequence[str],
    smiles_col: str = "SMILES alcohol",
    name_col: Optional[str] = "name",
    policy: str = "preferred",
    reference_col: Optional[str] = None,
    keep_cols: Sequence[str] = (),
) -> pd.DataFrame:
    """Esterify a library of alcohols with a single acid.

    alcohols       : DataFrame (uses `smiles_col`) or list of SMILES.
    policy         : "preferred" (1 product per alcohol) or "all" (regioisomers).
    reference_col  : column with reference SMILES to AUDIT (compares InChIKey and
                     flags match/mismatch).
    keep_cols      : columns from the input DataFrame to carry into the result.

    Returns a DataFrame with one row per product and validation columns.
    """
    if not isinstance(alcohols, pd.DataFrame):
        alcohols = pd.DataFrame({smiles_col: list(alcohols)})

    rows = []
    for _, r in alcohols.iterrows():
        alc_smi = str(r[smiles_col]).strip()
        name = str(r[name_col]) if (name_col and name_col in alcohols.columns) else None
        alc = Chem.MolFromSmiles(alc_smi)

        base = {"name": name, "alcohol_SMILES": alc_smi}
        for c in keep_cols:
            if c in alcohols.columns:
                base[c] = r[c]

        if alc is None:
            rows.append({**base, "SMILES": None, "status": "ERROR: alcohol did not parse"})
            continue

        ohs = classify_hydroxyls(alc)
        base["n_OH"] = len(ohs)
        base["OH_types"] = ", ".join(sorted({h["type"] for h in ohs}))

        prods = esterify(acid_smiles, alc_smi, policy=policy)
        if not prods:
            rows.append({**base, "SMILES": None, "status": "ERROR: reaction produced no product"})
            continue

        for p in prods:
            val = validate_ester(acid_smiles, alc_smi, p["smiles"])
            row = {
                **base,
                "SMILES": p["smiles"],
                "InChIKey": p["inchikey"],
                "esterified_OH": p["oh_type"],
                "fischer_viability": p["fischer_viability"],
                "n_regioisomers": p["n_regioisomers"],
                "formula": val["formula"],
                "MW": round(Descriptors.MolWt(Chem.MolFromSmiles(p["smiles"])), 2),
                "valid": val["valid"],
                "status": "OK" if val["valid"] else f"REVIEW: {val['reason']}",
            }
            # Audit against the previous reference, if requested.
            if reference_col and reference_col in alcohols.columns:
                ref = Chem.MolFromSmiles(str(r[reference_col]))
                ref_key = Chem.MolToInchiKey(ref) if ref is not None else None
                row["ref_SMILES"] = r[reference_col]
                if ref_key is None:
                    row["audit"] = "reference did not parse"
                elif ref_key == p["inchikey"]:
                    row["audit"] = "match"
                else:
                    same_skeleton = ref_key.split("-")[0] == p["inchikey"].split("-")[0]
                    row["audit"] = ("MISMATCH (stereochemistry)" if same_skeleton
                                    else "MISMATCH (skeleton/regiochemistry)")
            rows.append(row)

    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Name <-> structure consistency audit
# ---------------------------------------------------------------------------
def _name_variants(raw: str) -> list[str]:
    """Variants of a reagent name to try to resolve it.

    'Geraniol (2E)'                  -> ['Geraniol (2E)', 'Geraniol', '2E']
    'Prenol (3-methyl-2-buten-1-ol)' -> [..., '3-methyl-2-buten-1-ol']
    'beta-citronellol'               -> [..., 'citronellol']
    """
    import re
    n = str(raw or "").strip()
    if not n:
        return []
    out = [n]
    base = re.sub(r"\s*\([^)]*\)\s*$", "", n).strip()
    if base and base != n:
        out.append(base)
    m = re.search(r"\(([^)]*)\)\s*$", n)
    if m and m.group(1).strip():
        out.append(m.group(1).strip())
    no_prefix = re.sub(r"^(alpha|beta|gamma|n|sec|tert|cis|trans)[-\s]", "", base, flags=re.I)
    if no_prefix and no_prefix != base:
        out.append(no_prefix)
    return list(dict.fromkeys([x for x in out if x]))


def _resolve_name_to_smiles(name: str, use_opsin: bool = True,
                            use_pubchem: bool = True) -> tuple[Optional[str], Optional[str]]:
    """Resolve a NAME to SMILES. Returns (smiles, source)."""
    if use_opsin:
        try:
            from . import naming_smart as _ns
            if _ns.opsin_available():
                smi = _ns.opsin_to_smiles([name])[0]
                if smi:
                    return smi, "opsin"
        except Exception:
            pass
    if use_pubchem:
        try:
            import pubchempy as pcp
            comps = pcp.get_compounds(name, "name")
            if comps:
                c = comps[0]
                smi = (getattr(c, "isomeric_smiles", None)
                       or getattr(c, "smiles", None)
                       or getattr(c, "canonical_smiles", None))
                if smi:
                    return smi, "pubchem"
        except Exception:
            pass
    return None, None


def _has_defined_stereo(mol: Optional[Chem.Mol]) -> bool:
    """True if the molecule has ASSIGNED stereochemistry (centers or double bonds).

    Used to distinguish "the name does not specify stereochemistry" from "the
    name describes a different stereoisomer".
    """
    if mol is None:
        return False
    n_chiral = len(Chem.FindMolChiralCenters(mol, useLegacyImplementation=False,
                                             includeUnassigned=False))
    n_db = sum(1 for b in mol.GetBonds()
               if b.GetStereo() != Chem.BondStereo.STEREONONE)
    return (n_chiral + n_db) > 0


def audit_name_structure(
    df: pd.DataFrame,
    name_col: str,
    smiles_col: str,
    use_opsin: bool = True,
    use_pubchem: bool = True,
    pause: float = 0.30,
    progress: bool = True,
) -> pd.DataFrame:
    """Check that each row's NAME agrees with its SMILES.

    Resolves the name to structure (OPSIN -> PubChem) and compares InChIKey with
    the table SMILES. Detects the classic error of holding the wrong
    stereoisomer (e.g. writing "geraniol" with the SMILES of nerol), which is
    **invisible at a glance** because both share formula and flat drawing.

    verdict: 'match' | 'MISMATCH (stereochemistry)' |
             'MISMATCH (different structure)' | 'name not resolved'
    """
    import time
    rows = []
    it = df.iterrows()
    if progress:
        try:
            from tqdm.auto import tqdm
            it = tqdm(list(df.iterrows()), desc="Auditing name vs structure")
        except ImportError:
            pass

    for _, r in it:
        raw_name = r.get(name_col)
        smi = str(r.get(smiles_col, "")).strip()
        mol = Chem.MolFromSmiles(smi) if smi else None
        key = Chem.MolToInchiKey(mol) if mol is not None else None

        # Try ALL name variants: if any reproduces the SMILES it is a match
        # (avoids false positives from a variant the parser reads oddly).
        candidates = []          # [(smiles, source, inchikey)]
        for cand in _name_variants(raw_name):
            rsmi, rsrc = _resolve_name_to_smiles(
                cand, use_opsin=use_opsin, use_pubchem=use_pubchem)
            if rsmi:
                rmol_i = Chem.MolFromSmiles(rsmi)
                if rmol_i is not None:
                    rkey_i = Chem.MolToInchiKey(rmol_i)
                    candidates.append((rsmi, rsrc, rkey_i))
                    if key is not None and rkey_i == key:
                        break          # exact match: stop
            if pause:
                time.sleep(pause)

        exact = next((c for c in candidates if key is not None and c[2] == key), None)
        chosen = exact or (candidates[0] if candidates else (None, None, None))
        resolved_smi, resolved_src, rkey = chosen
        rmol = Chem.MolFromSmiles(resolved_smi) if resolved_smi else None

        if key is None:
            verdict = "SMILES did not parse"
        elif rkey is None:
            verdict = "name not resolved"
        elif rkey == key:
            verdict = "match"
        elif rkey.split("-")[0] == key.split("-")[0]:
            # Same connectivity: tell a REAL mismatch from an ambiguous name.
            if not _has_defined_stereo(rmol):
                verdict = "match (name has no defined stereochemistry)"
            elif not _has_defined_stereo(mol):
                verdict = "REVIEW: your SMILES has no defined stereochemistry"
            else:
                verdict = "MISMATCH (stereochemistry)"
        else:
            verdict = "MISMATCH (different structure)"

        rows.append({
            "name": raw_name,
            "table_SMILES": smi,
            "SMILES_from_name": resolved_smi,
            "resolved_by": resolved_src,
            "verdict": verdict,
        })
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Loading (with cell colors)
# ---------------------------------------------------------------------------
def load_alcohols(path: str, sheet: str | int = 0) -> pd.DataFrame:
    """Load a table of alcohols from .xlsx or .csv."""
    if str(path).lower().endswith((".xlsx", ".xlsm", ".xls")):
        return pd.read_excel(path, sheet_name=sheet)
    return pd.read_csv(path)


# Fill-color -> category map (legend of the benzofuroxane workbook).
COLOR_CATEGORIES: dict[str, str] = {
    "FFFFFF": "existing (previous list)",
    "FFF9C4": "corrected SMILES (check E/Z)",
    "FFF3CD": "synthetic flag (tertiary/phenolic OH)",
    "D6EAF8": "inventory (available in lab)",
    "FDECEA": "approximate SMILES (verify)",
}


def read_fill_categories(
    path: str,
    sheet: str | int = 0,
    key_column: str = "name",
    color_map: Optional[dict] = None,
) -> dict[str, str]:
    """Read each row's FILL COLOR and translate it into a category.

    Many lab spreadsheets encode information in colors (e.g. which reagents are
    in inventory) that `pandas` does not see. Returns
    {key_column value: category}.
    """
    color_map = color_map or COLOR_CATEGORIES
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {}
    try:
        wb = load_workbook(path)
        ws = wb[sheet] if isinstance(sheet, str) else wb.worksheets[sheet]
        header = [c.value for c in ws[1]]
        if key_column not in header:
            return {}
        ci = header.index(key_column)
        out: dict[str, str] = {}
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            cell = row[ci]
            if cell.value is None:
                continue
            rgb = getattr(cell.fill.start_color, "rgb", None) if cell.fill else None
            rgb = rgb[-6:].upper() if isinstance(rgb, str) else None
            out[str(cell.value)] = color_map.get(rgb, "no category")
        return out
    except Exception:
        return {}


def load_alcohols_with_categories(
    path: str,
    sheet: str | int = 0,
    key_column: str = "name",
) -> pd.DataFrame:
    """Load the table and add a `category` column inferred from the color."""
    df = load_alcohols(path, sheet=sheet)
    cats = read_fill_categories(path, sheet=sheet, key_column=key_column)
    if cats and key_column in df.columns:
        df["category"] = df[key_column].astype(str).map(cats).fillna("no category")
    return df
